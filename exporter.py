"""数据导出：JSON / CSV / Excel。"""
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class Exporter:
    @staticmethod
    def to_json(data, filepath):
        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    @staticmethod
    def to_csv(rows, filepath, columns=None):
        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        if not rows:
            with open(path, "w", encoding="utf-8") as f:
                f.write("")
            return
        columns = columns or list(rows[0].keys())
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

    @staticmethod
    def to_excel(results, filepath):
        """生成多 Sheet 的 Excel 文件。"""
        wb = Workbook()
        wb.remove(wb.active)  # 移除默认空白 sheet

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        cell_align = Alignment(vertical="center")

        def write_sheet(title, columns, rows):
            ws = wb.create_sheet(title=title)
            # 写表头
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            # 写数据行
            for row_idx, row in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    val = row.get(col_name, "")
                    if col_name in ("actual_hours", "planned_hours") and val == 0:
                        val = ""
                    if isinstance(val, bool):
                        val = "是" if col_name in ("is_done", "is_archived") and val else "否"
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = cell_align
            # 设置列宽
            col_widths = {
                "project_id": 26, "project_name": 28, "project_description": 36,
                "stage_id": 26, "stage_name": 28,
                "task_id": 26, "content": 48, "executor_id": 26,
                "is_done": 10, "is_archived": 10, "task_count": 10,
                "actual_hours": 14, "planned_hours": 14,
            }
            for col_idx, col_name in enumerate(columns, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
                    col_widths.get(col_name, 16)

        # Sheet 1: 项目清单
        projects = results.get("projects", [])
        write_sheet("项目清单",
                    ["project_name", "project_id", "project_description", "is_archived",
                     "task_count", "actual_hours", "planned_hours"],
                    projects)

        # Sheet 2: 任务列表
        stages = results.get("stages", [])
        write_sheet("任务列表",
                    ["stage_name", "stage_id", "project_name"],
                    stages)

        # Sheet 3: 任务
        tasks = results.get("tasks", [])
        write_sheet("任务",
                    ["content", "project_name", "stage_name", "is_done", "executor_id",
                     "actual_hours", "planned_hours"],
                    tasks)

        # Sheet 4: 工时（仅含工时记录的任务）
        hours = [t for t in tasks if t.get("actual_hours", 0) > 0 or t.get("planned_hours", 0) > 0]
        write_sheet("工时",
                    ["content", "project_name", "stage_name", "actual_hours", "planned_hours"],
                    hours)

        wb.save(filepath)
